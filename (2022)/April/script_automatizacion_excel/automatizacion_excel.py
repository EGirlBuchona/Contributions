# Gustavo Cocone Zavaleta | GitHub: @EgirlBuchona | © 2022
# Proyecto: Automatización de Archivos Excel en Python

import openpyxl
from openpyxl.styles import Font, Alignment

def crear_archivo_excel(nombre_archivo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws["A1"] = "ID"
    ws["B1"] = "Nombre"
    ws["C1"] = "Edad"
    ws["A1"].font = ws["B1"].font = ws["C1"].font = Font(bold=True)
    ws["A1"].alignment = ws["B1"].alignment = ws["C1"].alignment = Alignment(horizontal="center")
    wb.save(nombre_archivo)
    print(f"Archivo '{nombre_archivo}' creado exitosamente.")

def agregar_datos(nombre_archivo, datos):
    wb = openpyxl.load_workbook(nombre_archivo)
    ws = wb.active
    for fila, (id, nombre, edad) in enumerate(datos, start=2):
        ws[f"A{fila}"] = id
        ws[f"B{fila}"] = nombre
        ws[f"C{fila}"] = edad
    wb.save(nombre_archivo)
    print("Datos agregados exitosamente.")

def leer_datos(nombre_archivo):
    wb = openpyxl.load_workbook(nombre_archivo)
    ws = wb.active
    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        print(fila)

if __name__ == "__main__":
    archivo = "datos_usuarios.xlsx"
    crear_archivo_excel(archivo)

    datos = [
        (1, "Juan Pérez", 30),
        (2, "Ana López", 25),
        (3, "Carlos Gómez", 35)
    ]
    agregar_datos(archivo, datos)
    print("Contenido del archivo:")
    leer_datos(archivo)
